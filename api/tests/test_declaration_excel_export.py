from io import BytesIO

from django.urls import reverse

from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from data.factories import InstructionRoleFactory, OngoingInstructionDeclarationFactory

from .utils import authenticate


class TestDeclarationExcelExport(APITestCase):
    def test_excel_export_unauthenticated(self):
        response = self.client.get(reverse("api:export_excel_declarations"))
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @authenticate
    def test_excel_export_response(self):
        InstructionRoleFactory(user=authenticate.user)
        d1 = OngoingInstructionDeclarationFactory()
        d2 = OngoingInstructionDeclarationFactory()

        response = self.client.get(reverse("api:export_excel_declarations"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # Est-ce que le contenu est un fichier Excel ?
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response["Content-Type"])

        # Vérifier les contenus du ficher
        workbook = load_workbook(filename=BytesIO(response.content))
        worksheet = workbook.active

        # Vérification de l'entête
        headers = [cell.value for cell in worksheet[1]]
        expected_headers = (
            "Nom du produit",
            "Marque",
            "Article",
            "Statut",
            "Nom de l'entreprise",
            "No. SIRET",
            "No. TVA",
            "No. de département",
        )
        for h in expected_headers:
            self.assertIn(h, headers)

        # Vérification du contenu
        data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        self.assertGreaterEqual(len(data_rows), 2)
        self.assertIn(d1.name, [row[headers.index("Nom du produit")] for row in data_rows])
        self.assertIn(d2.name, [row[headers.index("Nom du produit")] for row in data_rows])
