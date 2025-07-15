import logging
from io import BytesIO

from django.db.models import Case, DateTimeField, F, OuterRef, Q, Subquery, Value, When
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404

from django_filters import rest_framework as django_filters
from drf_excel.mixins import XLSXFileMixin
from drf_excel.renderers import XLSXRenderer
from openpyxl import load_workbook
from rest_framework.exceptions import NotFound, PermissionDenied
from rest_framework.generics import (
    GenericAPIView,
    ListAPIView,
    ListCreateAPIView,
    RetrieveAPIView,
    RetrieveUpdateDestroyAPIView,
)
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.response import Response
from viewflow.fsm.base import TransitionNotAllowed

from api.exceptions import ProjectAPIException
from api.permissions import (
    CanAccessIndividualDeclaration,
    CanAccessUserDeclatarions,
    CanTakeAuthorship,
    IsController,
    IsDeclarant,
    IsDeclarationAuthor,
    IsInstructor,
    IsSupervisor,
    IsVisor,
)
from api.serializers import (
    ControllerDeclarationSerializer,
    DeclarationSerializer,
    DeclarationShortSerializer,
    ExcelExportDeclarationSerializer,
    OpenDataDeclarationSerializer,
    SimpleDeclarationSerializer,
    SimpleInstructorSerializer,
    SimpleUserSerializer,
    SimpleVisorSerializer,
)
from api.utils.filters import CamelCaseOrderingFilter, SimplifiedStatusFilter
from api.utils.search import UnaccentSearchFilter
from api.views.declaration.declaration_filter_set import DeclarationFilterSet
from api.views.declaration.declaration_flow import DeclarationFlow
from config import email
from data.models import Company, Declaration, InstructionRole, Snapshot, User, VisaRole

logger = logging.getLogger(__name__)


class DeclarationPagination(LimitOffsetPagination):
    default_limit = 10
    max_limit = 50


class UserDeclarationPagination(DeclarationPagination):
    """
    On ajoute dans le payload les instructrices et viseuses assignées au dossiers afin de pouvoir
    filtrer par personne assignée.
    """

    authors = []

    def paginate_queryset(self, queryset, request, view=None):
        original_queryset = view.get_queryset()
        authors_ids = original_queryset.values_list("author", flat=True).distinct().order_by()
        self.authors = (User.objects.get(pk=x) for x in authors_ids if User.objects.filter(pk=x).exists())
        return super().paginate_queryset(queryset, request, view)

    def get_paginated_response(self, data):
        return Response(
            {
                "count": self.count,
                "next": self.get_next_link(),
                "previous": self.get_previous_link(),
                "results": data,
                "authors": SimpleUserSerializer(self.authors, many=True).data,
            }
        )


class UserDeclarationsListCreateApiView(ListCreateAPIView):
    model = Declaration
    permission_classes = [CanAccessUserDeclatarions]
    search_fields = ["name", "id", "company__social_name", "teleicare_declaration_number"]
    pagination_class = UserDeclarationPagination
    filter_backends = [
        django_filters.DjangoFilterBackend,
        CamelCaseOrderingFilter,
        UnaccentSearchFilter,
    ]
    filterset_class = DeclarationFilterSet

    def get_queryset(self):
        declarable_companies = self.request.user.declarable_companies.all()
        all_companies = list(declarable_companies.union(self.request.user.supervisable_companies.all()))
        return Declaration.objects.filter(
            Q(author=self.request.user) | Q(company__in=all_companies) | Q(mandated_company__in=all_companies)
        )

    def perform_create(self, serializer):
        # Lors de la création, des validations concernant l'objet créé doivent être faits ici
        # https://www.django-rest-framework.org/api-guide/permissions/#limitations-of-object-level-permissions
        company_id = self.request.data.get("company")
        try:
            company = Company.objects.get(pk=company_id)
        except Company.DoesNotExist as _:
            raise NotFound("Company not found")

        companies = [company] + list(company.mandated_companies.all())
        if not any(x for x in companies if x.declarant_roles.filter(user=self.request.user).exists()):
            raise PermissionDenied()
        return super().perform_create(serializer)

    def get_serializer_class(self):
        if self.request.method == "POST":
            return DeclarationSerializer
        return DeclarationShortSerializer


class DeclarationRetrieveUpdateDestroyView(RetrieveUpdateDestroyAPIView):
    model = Declaration
    serializer_class = DeclarationSerializer
    permission_classes = [CanAccessIndividualDeclaration]

    def get_queryset(self):
        queryset = Declaration.objects.all()
        queryset = self.get_serializer_class().setup_eager_loading(queryset)
        return queryset

    def perform_destroy(self, instance):
        if instance.status == Declaration.DeclarationStatus.DRAFT:
            return super().perform_destroy(instance)
        raise ProjectAPIException(global_error="Les déclarations en cours ne peuvent pas être supprimées.")


class DeclarationPagination(LimitOffsetPagination):
    default_limit = 10
    max_limit = 50


class InstructionDeclarationPagination(DeclarationPagination):
    """
    On ajoute dans le payload les instructrices et viseuses assignées au dossiers afin de pouvoir
    filtrer par personne assignée.
    """

    instructors = []
    visors = []

    def paginate_queryset(self, queryset, request, view=None):
        self.instructors = InstructionRole.objects.filter(id__in=InstructionRole.objects.values_list("id"))
        self.visors = VisaRole.objects.filter(id__in=VisaRole.objects.values_list("id"))
        return super().paginate_queryset(queryset, request, view)

    def get_paginated_response(self, data):
        return Response(
            {
                "count": self.count,
                "next": self.get_next_link(),
                "previous": self.get_previous_link(),
                "results": data,
                "instructors": SimpleInstructorSerializer(self.instructors, many=True).data,
                "visors": SimpleVisorSerializer(self.visors, many=True).data,
            }
        )


class InstructionDateOrderingFilter(CamelCaseOrderingFilter):
    """
    Pour filtrer par la date limite d'instruction on doit faire un filtre custom car cette date appartient
    à un autre modèle : le Snapshot.
    Même si on a la property dans le modèle Declaration, on ne peut pas l'utiliser directement car cette
    property n'est pas en base de données, elle est déjà dans la couche Django/Python. Pour l'avoir dans la
    couche DB on doit faire une annotation qui la contient.
    """

    def get_ordering(self, request, queryset, view):
        ordering = super().get_ordering(request, queryset, view)
        fields_to_remove = ("response_limit_date", "-response_limit_date")
        for field in fields_to_remove:
            try:
                ordering.remove(field)
            except Exception as _:
                pass
        return ordering

    def order_by_response_limit(self, request):
        """
        Returns tupple :
        (effectuer un triage par date d'instruction, desc)
        """
        return (
            "responseLimitDate" in request.query_params.get(self.ordering_param, ""),
            "-responseLimitDate" in request.query_params.get(self.ordering_param, ""),
        )

    def filter_queryset(self, request, queryset, view):
        """
        Cette fonction vise à réproduire la property "response_limit_date" du modèle Declaration
        mais dans la couche DB (avec des querysets) afin de pouvoir filtrer dessus.

        ⚠️ Attention : Tout changement effectué dans cette fonction doit aussi être reflété dans
        data/models/declaration.py > response_limit_date
        """
        order_by_response_limit, desc = self.order_by_response_limit(request)

        qs = super().filter_queryset(request, queryset, view)
        if not order_by_response_limit:
            return qs

        latest_snapshot_subquery = (
            Snapshot.objects.filter(
                declaration=OuterRef("pk"), status=Declaration.DeclarationStatus.AWAITING_INSTRUCTION
            )
            .exclude(action=Snapshot.SnapshotActions.REFUSE_VISA)
            .order_by("-creation_date" if order_by_response_limit else "creation_date")
            .values("creation_date")[:1]
        )

        concerned_statuses = [
            Declaration.DeclarationStatus.AWAITING_INSTRUCTION,
            Declaration.DeclarationStatus.ONGOING_INSTRUCTION,
            Declaration.DeclarationStatus.AWAITING_VISA,
            Declaration.DeclarationStatus.ONGOING_VISA,
        ]

        # On met toujours les valeurs `None` à la fin pour ne pas devoir changer de page
        # pour voir les premières déclarations ayant la date remplie
        order_function = (
            F("annotated_response_limit_date").desc(nulls_last=True) if desc else F("annotated_response_limit_date")
        )
        qs = qs.annotate(
            annotated_response_limit_date=Case(
                When(
                    status__in=concerned_statuses,
                    then=Coalesce(Subquery(latest_snapshot_subquery, output_field=DateTimeField()), Value(None)),
                ),
                default=Value(None),
                output_field=DateTimeField(),
            )
        ).order_by(order_function)

        return qs


class GenericDeclarationsListView(ListAPIView):
    model = Declaration
    pagination_class = DeclarationPagination
    filter_backends = [
        django_filters.DjangoFilterBackend,
        CamelCaseOrderingFilter,
    ]
    filterset_class = DeclarationFilterSet


class CommonOngoingDeclarationView(GenericDeclarationsListView):
    permission_classes = [(IsInstructor | IsVisor)]
    search_fields = ["name", "id", "company__social_name", "teleicare_declaration_number"]
    filter_backends = [
        django_filters.DjangoFilterBackend,
        InstructionDateOrderingFilter,
        UnaccentSearchFilter,
    ]
    ordering_fields = ["creation_date", "modification_date", "name", "response_limit_date"]
    queryset = Declaration.objects.exclude(status=Declaration.DeclarationStatus.DRAFT).distinct()


class DeclarationHyperlinkXLSXRenderer(XLSXRenderer):
    """
    Ceci est un hack pour ajouter des liens au noms de la déclaration. Cet issue
    a été levé dans drf-excel : https://github.com/django-commons/drf-excel/issues/112
    """

    def render(self, data, media_type=None, renderer_context=None):
        excel_data = super().render(data, media_type, renderer_context)
        workbook = load_workbook(filename=BytesIO(excel_data))

        # Ajouter le lien vers le résultat de la recherche dans le nom du produit
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=2):  # L'index commence à 1, et on veux éviter l'entête, donc 2
                name_cell = row[0]
                hyperlink_data = data[name_cell.row - 2]["url_field"]
                name_cell.value = hyperlink_data.display
                name_cell.hyperlink = hyperlink_data.ref

        # Sauvegarder en Bytes
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


class OngoingDeclarationsExcelView(XLSXFileMixin, CommonOngoingDeclarationView):
    pagination_class = None
    renderer_classes = [DeclarationHyperlinkXLSXRenderer]
    serializer_class = ExcelExportDeclarationSerializer
    filename = "declarations-resultats.xlsx"

    max_rows = 5000

    def filter_queryset(self, queryset):
        """
        Permet de retourner un maximum de max_rows déclarations
        """
        queryset = super().filter_queryset(queryset)
        return queryset[: self.max_rows]

    # Format de l'entête du fichier Excel
    column_header = {
        "titles": [
            "Nom du produit",
            "Marque",
            "Article",
            "Statut",
            "Nom de l'entreprise",
            "No. SIRET",
            "No. TVA",
            "No. de département",
        ],
        "column_width": [30, 20, 30, 25, 30, 15, 15, 15],
        "height": 30,
        "style": {
            "fill": {
                "fill_type": "solid",
                "start_color": "FF000091",
            },
            "alignment": {
                "horizontal": "center",
                "vertical": "center",
                "wrapText": True,
                "shrink_to_fit": True,
            },
            "border_side": {
                "border_style": "thin",
                "color": "FF6A6AF4",
            },
            "font": {
                "name": "Arial",
                "size": 12,
                "bold": True,
                "color": "FFFFFFFF",
            },
        },
    }
    body = {"height": 20}


class OngoingDeclarationsListView(CommonOngoingDeclarationView):
    pagination_class = InstructionDeclarationPagination
    serializer_class = SimpleDeclarationSerializer


class ControllerDeclarationsListView(CommonOngoingDeclarationView):
    permission_classes = [IsController]
    pagination_class = DeclarationPagination
    ordering_fields = ["name", "company_name"]
    serializer_class = ControllerDeclarationSerializer
    filter_backends = [
        SimplifiedStatusFilter,
        django_filters.DjangoFilterBackend,
        CamelCaseOrderingFilter,
        UnaccentSearchFilter,
    ]

    def get_queryset(self):
        queryset = super().get_queryset().select_related("company")
        return queryset.annotate(company_name=F("company__social_name"))


class ControllerDeclarationRetrieveView(RetrieveAPIView):
    permission_classes = [IsController]
    model = Declaration
    serializer_class = DeclarationSerializer
    queryset = Declaration.objects.exclude(status=Declaration.DeclarationStatus.DRAFT).distinct()


class OpenDataDeclarationsListView(GenericDeclarationsListView):
    serializer_class = OpenDataDeclarationSerializer
    filter_backends = [django_filters.DjangoFilterBackend]
    ordering_fields = ["creation_date", "modification_date", "name", "response_limit_date"]

    def get_queryset(self):
        queryset = Declaration.objects.filter(status=Declaration.DeclarationStatus.AUTHORIZED)
        queryset = self.get_serializer_class().setup_eager_loading(queryset)
        return queryset


class CompanyDeclarationsListView(GenericDeclarationsListView):
    # Une fois l'instruction commencée on pourra avoir un serializer différent pour cette vue
    serializer_class = SimpleDeclarationSerializer
    permission_classes = [IsSupervisor]

    def get_queryset(self):
        company = get_object_or_404(
            Company.objects.filter(supervisors=self.request.user, pk=self.kwargs[self.lookup_field])
        )
        return company.declarations.exclude(status=Declaration.DeclarationStatus.DRAFT)


class DeclarationTakeAuthorshipView(GenericAPIView):
    serializer_class = SimpleDeclarationSerializer
    permission_classes = [CanTakeAuthorship]
    queryset = Declaration.objects.all()

    def post(self, request, pk):
        declaration = self.get_object()

        declaration.author = request.user
        declaration.save()
        declaration.refresh_from_db()
        serializer = self.get_serializer(declaration)
        return Response(serializer.data)


class DeclarationAssignInstruction(GenericAPIView):
    """
    Contrairement à `DeclarationTakeForInstructionView`, cette view ne change pas
    le statut de la déclaration, elle sert simplement à transférer l'instruction d'un
    dossier à l'instrutrice faisant la requête
    """

    serializer_class = SimpleDeclarationSerializer
    permission_classes = [IsInstructor]
    queryset = Declaration.objects.all()

    def post(self, request, pk):
        declaration = self.get_object()

        declaration.instructor = request.user.instructionrole
        declaration.save()
        declaration.refresh_from_db()
        serializer = self.get_serializer(declaration)
        return Response(serializer.data)


class ArticleChangeView(GenericAPIView):
    permission_classes = [(IsInstructor | IsVisor)]
    serializer_class = DeclarationSerializer

    def get_queryset(self):
        queryset = Declaration.objects.all()
        queryset = self.get_serializer_class().setup_eager_loading(queryset)
        return queryset

    def post(self, request, pk):
        declaration = self.get_object()
        new_article = request.data.get("article", "")

        if new_article not in Declaration.Article:
            raise ProjectAPIException(global_error="Merci de spécifier un article valide")

        declaration.overridden_article = Declaration.Article(new_article)
        declaration.save()
        declaration.refresh_from_db()
        serializer = self.get_serializer(declaration)
        return Response(serializer.data)


class DeclarationFlowView(GenericAPIView):
    serializer_class = DeclarationSerializer
    transition = None
    create_snapshot = False
    snapshot_action = Snapshot.SnapshotActions.OTHER
    snapshot_post_validation_status = ""
    from_status = None
    to_status = None
    brevo_template_id = None  # Remplir avec l'ID du template Brevo si un email doit être envoyé

    def get_queryset(self):
        queryset = Declaration.objects.all()
        queryset = self.get_serializer_class().setup_eager_loading(queryset)
        return queryset

    def get_brevo_template_id(self, request, declaration):
        """
        À surcharger en cas du besoin d'un template email dynamique
        """
        return self.brevo_template_id

    def get_brevo_parameters(self, request, declaration):
        """
        À surcharger en cas des paramètres spéciaux à envoyer à Brevo
        """
        return declaration.brevo_parameters

    def get_transition(self, request, declaration):
        """
        À surcharger en cas d'une transition dynamique
        """
        return self.transition

    def get_snapshot_action(self, request, declaration):
        """
        À surcharger en cas d'une action dynamique
        """
        return self.snapshot_action

    def get_snapshot_post_validation_status(self, request, declaration):
        """
        À surcharger en cas d'un statut dynamique
        """
        return self.snapshot_post_validation_status

    def on_transition_success(self, request, declaration):
        """
        Hook à surcharger dans le cas où il faut faire des choses particulières
        après le succès de la transition
        """
        pass

    def perform_snapshot_creation(self, request, declaration):
        """
        Possible de le surcharger si la création du snapshot nécessite un
        traitement spécial
        """
        declaration.create_snapshot(
            user=request.user,
            comment=request.data.get("comment", ""),
            expiration_days=request.data.get("expiration"),
            action=self.get_snapshot_action(request, declaration),
            post_validation_status=self.get_snapshot_post_validation_status(request, declaration),
            blocking_reasons=request.data.get("reasons"),
        )

    def post(self, request, *args, **kwargs):
        declaration = self.get_object()
        self.from_status = declaration.status
        flow = DeclarationFlow(declaration)
        transition_method = getattr(flow, self.get_transition(request, declaration))
        flow_permission_method = getattr(transition_method, "has_permission", None)
        if flow_permission_method and not flow_permission_method(request.user):
            raise PermissionDenied()
        try:
            transition_method()
        except TransitionNotAllowed as e:
            logger.exception(e)
            raise ProjectAPIException(global_error="Erreur lors du changement de statut de la déclaration")
        self.to_status = declaration.status
        brevo_template_id = self.get_brevo_template_id(request, declaration)
        if self.create_snapshot:
            self.perform_snapshot_creation(request, declaration)
        declaration.assign_calculated_article()
        self.on_transition_success(request, declaration)
        if brevo_template_id and declaration.author:
            try:
                email.send_sib_template(
                    brevo_template_id,
                    self.get_brevo_parameters(request, declaration),
                    declaration.author.email,
                    declaration.author.get_full_name(),
                )
            except Exception as e:
                logger.error(f"Email not sent on transition {self.get_transition(request, declaration)}")
                logger.exception(e)
        declaration.save()
        serializer = self.get_serializer(declaration)
        return Response(serializer.data)


class DeclarationSubmitView(DeclarationFlowView):
    """
    DRAFT -> AWAITING_INSTRUCTION
    """

    permission_classes = [IsDeclarationAuthor, IsDeclarant]
    transition = "submit"
    create_snapshot = True
    snapshot_action = Snapshot.SnapshotActions.SUBMIT
    brevo_template_id = 3


class DeclarationTakeForInstructionView(DeclarationFlowView):
    """
    AWAITING_INSTRUCTION -> ONGOING_INSTRUCTION
    """

    permission_classes = [IsInstructor]
    transition = "take_for_instruction"
    create_snapshot = True
    snapshot_action = Snapshot.SnapshotActions.TAKE_FOR_INSTRUCTION

    def on_transition_success(self, request, declaration):
        declaration.instructor = InstructionRole.objects.get(user=request.user)
        return super().on_transition_success(request, declaration)


class DeclarationTakeForVisaView(DeclarationFlowView):
    """
    AWAITING_VISA -> ONGOING_VISA
    """

    permission_classes = [IsVisor]
    transition = "take_for_visa"
    create_snapshot = True
    snapshot_action = Snapshot.SnapshotActions.TAKE_FOR_VISA

    def on_transition_success(self, request, declaration):
        declaration.visor = VisaRole.objects.get(user=request.user)
        return super().on_transition_success(request, declaration)


class DeclarationObserveView(DeclarationFlowView):
    """
    ONGOING_INSTRUCTION -> OBSERVATION
    """

    permission_classes = [IsInstructor]
    transition = "observe_no_visa"
    create_snapshot = True
    snapshot_action = Snapshot.SnapshotActions.OBSERVE_NO_VISA
    brevo_template_id = 4

    def on_transition_success(self, request, declaration):
        declaration.instructor = InstructionRole.objects.get(user=request.user)
        return super().on_transition_success(request, declaration)


class DeclarationAuthorizeView(DeclarationFlowView):
    """
    ONGOING_INSTRUCTION -> AUTHORIZED
    """

    permission_classes = [IsInstructor]
    transition = "authorize_no_visa"
    create_snapshot = True
    snapshot_action = Snapshot.SnapshotActions.AUTHORIZE_NO_VISA
    brevo_template_id = 6

    def on_transition_success(self, request, declaration):
        declaration.instructor = InstructionRole.objects.get(user=request.user)
        return super().on_transition_success(request, declaration)


class DeclarationResubmitView(DeclarationFlowView):
    """
    [OBSERVATION, OBJECTION] -> ONGOING_INSTRUCTION
    """

    permission_classes = [IsDeclarationAuthor, IsDeclarant]
    transition = "resubmit"
    create_snapshot = True

    def get_snapshot_action(self, request, declaration):
        if self.from_status == Declaration.DeclarationStatus.OBSERVATION:
            return Snapshot.SnapshotActions.RESPOND_TO_OBSERVATION
        return Snapshot.SnapshotActions.RESPOND_TO_OBJECTION


class DeclarationWithdrawView(DeclarationFlowView):
    """
    AUTHORIZED -> WITHDRAWN
    """

    permission_classes = [IsDeclarationAuthor | IsDeclarant]
    transition = "withdraw"
    create_snapshot = True
    snapshot_action = Snapshot.SnapshotActions.WITHDRAW
    brevo_template_id = 8

    def perform_snapshot_creation(self, request, declaration):
        """
        Il faut renseigner la date effective de retrait du marché
        """
        declaration.create_snapshot(
            user=request.user,
            action=self.get_snapshot_action(request, declaration),
            effective_withdrawal_date=request.data.get("effective_withdrawal_date"),
        )


class DeclarationAbandonView(DeclarationFlowView):
    permission_classes = [(IsDeclarationAuthor | IsDeclarant)]
    transition = "abandon"
    create_snapshot = True
    snapshot_action = Snapshot.SnapshotActions.ABANDON


class VisaDecisionView(DeclarationFlowView):
    permission_classes = [IsVisor]
    create_snapshot = True

    def on_transition_success(self, request, declaration):
        """
        Dans le cas d'un refus de visa, on remet à vide les champs post_validation.
        On met également les notes privées à destination de l'admnistration dans la déclaration.
        """
        declaration.post_validation_producer_message = ""
        declaration.post_validation_status = ""
        declaration.post_validation_expiration_days = None
        return super().on_transition_success(request, declaration)


class DeclarationRefuseVisaView(VisaDecisionView):
    """
    ONGOING_VISA -> AWAITING_INSTRUCTION
    """

    transition = "refuse_visa"
    snapshot_action = Snapshot.SnapshotActions.REFUSE_VISA

    def get_snapshot_post_validation_status(self, request, declaration):
        return declaration.post_validation_status

    def perform_snapshot_creation(self, request, declaration):
        declaration.create_snapshot(
            user=request.user,
            action=self.get_snapshot_action(request, declaration),
            post_validation_status=self.get_snapshot_post_validation_status(request, declaration),
        )


class DeclarationAcceptVisaView(VisaDecisionView):
    """
    ONGOING_VISA -> { AUTHORIZED | REJECTED | OBJECTION | OBSERVATION }
    Le ou la viseuse peut surcharger la décision de l'instructrice en envoyant
    un objet dans le payload de cette forme :
    {
        comment: "overridden comment",
        proposal: "OBSERVATION", // (ou un autre statut)
        delayDays: 10,
        reasons: ["a", "b"],
    }
    """

    snapshot_action = Snapshot.SnapshotActions.ACCEPT_VISA

    def get_validation_status(self, request, declaration):
        overridden_status = request.data.get("proposal")
        if overridden_status:
            return Declaration.DeclarationStatus(overridden_status)
        return declaration.post_validation_status

    def get_snapshot_post_validation_status(self, request, declaration):
        return self.get_validation_status(request, declaration)

    def perform_snapshot_creation(self, request, declaration):
        """
        Possible de le surcharger si la création du snapshot nécessite un
        traitement spécial
        """
        overridden = request.data.get("proposal")
        data = request.data
        d = declaration
        d.create_snapshot(
            user=request.user,
            comment=data.get("comment", "") if overridden else d.post_validation_producer_message,
            expiration_days=data.get("delay_days") if overridden else d.post_validation_expiration_days,
            action=self.get_snapshot_action(request, d),
            post_validation_status=self.get_snapshot_post_validation_status(request, d),
            blocking_reasons=data.get("reasons") if overridden else None,
        )

    def get_transition(self, request, declaration):
        transition_map = {
            Declaration.DeclarationStatus.AUTHORIZED: "accept_visa_authorize",
            Declaration.DeclarationStatus.REJECTED: "accept_visa_reject",
            Declaration.DeclarationStatus.OBJECTION: "accept_visa_object",
            Declaration.DeclarationStatus.OBSERVATION: "accept_visa_observe",
        }
        return transition_map.get(self.get_validation_status(request, declaration))

    def get_brevo_template_id(self, request, declaration):
        template_map = {
            Declaration.DeclarationStatus.AUTHORIZED: 6,
            Declaration.DeclarationStatus.REJECTED: 7,
            Declaration.DeclarationStatus.OBJECTION: 5,
            Declaration.DeclarationStatus.OBSERVATION: 4,
        }
        return template_map.get(self.get_validation_status(request, declaration))


# Nous utilisons une Non-deterministic state machine. Lors qu'un.e instructeur.ice demande une
# visa, nous assignnos la déclaration à `AWAITING_VISA` et ajoutons un propriété dans le modèle
# spécifiant quel sera le status à assigner après le flow de validation.
class VisaRequestFlowView(DeclarationFlowView):
    """
    ONGOING_INSTRUCTION -> AWAITING_VISA
    Cette view doit être sous-classée. Elle assigne le `post_validation_status`
    spécifié. Les sous-classes doivent donc déclarer cette propriété.
    """

    permission_classes = [IsInstructor]
    transition = "request_visa"
    post_validation_status = None
    create_snapshot = True
    snapshot_action = Snapshot.SnapshotActions.REQUEST_VISA

    def perform_snapshot_creation(self, request, declaration):
        """
        Dans le cas d'une requête de validation, on ne mettra pas le commentaire à
        destination du producteur dans le snapshot créé. On le met dans le modèle pour
        pouvoir l'envoyer au producteur si la décision est acceptée par la viseuse.
        On met également les notes privées à destination de l'admnistration dans la déclaration.
        """
        declaration.create_snapshot(
            user=request.user,
            action=self.get_snapshot_action(request, declaration),
            post_validation_status=self.post_validation_status,
            blocking_reasons=request.data.get("reasons"),
        )

    def on_transition_success(self, request, declaration):
        declaration.post_validation_producer_message = request.data.get("comment", "")
        declaration.post_validation_expiration_days = request.data.get("expiration")
        declaration.instructor = InstructionRole.objects.get(user=request.user)

        if not self.post_validation_status:
            raise Exception("VisaRequestFlowView doit être sous-classée et doit spécifier le post_validation_status")
        declaration.post_validation_status = self.post_validation_status
        return super().on_transition_success(request, declaration)


class DeclarationObserveWithVisa(VisaRequestFlowView):
    post_validation_status = Declaration.DeclarationStatus.OBSERVATION


class DeclarationObjectWithVisa(VisaRequestFlowView):
    post_validation_status = Declaration.DeclarationStatus.OBJECTION


class DeclarationRejectWithVisa(VisaRequestFlowView):
    post_validation_status = Declaration.DeclarationStatus.REJECTED


class DeclarationAuthorizeWithVisa(VisaRequestFlowView):
    post_validation_status = Declaration.DeclarationStatus.AUTHORIZED
